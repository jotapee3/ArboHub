from __future__ import annotations

import tempfile
import unittest
from datetime import date
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from app.services.qualifica.relatorio_60d_service import (
    Notificacao60d,
    Relatorio60dService,
)
from app.services.qualifica.relatorio_72h_service import MunicipioQualifica


class Relatorio60dServiceTestCase(unittest.TestCase):
    def setUp(self):
        self.service = Relatorio60dService()
        self.municipios = (
            MunicipioQualifica("430003", "Aceguá", 7),
            MunicipioQualifica("430005", "Água Santa", 6),
            MunicipioQualifica("431490", "Porto Alegre", 2),
        )

    def test_classifica_todos_os_motivos_de_nao_oportunidade(self):
        notificacoes = (
            self._notificacao("1", "430003", dias_encerramento=0),
            self._notificacao("1", "430003", dias_encerramento=60),
            self._notificacao("3", "430003", dias_encerramento=61),
            self._notificacao("4", "430003", dias_encerramento=None),
            self._notificacao("5", "430003", dias_encerramento=-1),
            self._notificacao(
                "6",
                "430003",
                dias_encerramento=10,
                classificacao=8,
            ),
            self._notificacao(
                "7",
                "430003",
                dias_encerramento=10,
                classificacao=99,
            ),
            self._notificacao(
                "8",
                "430003",
                dias_encerramento=30,
                classificacao=12,
            ),
            self._notificacao(
                "sentinela",
                "430003",
                municipio_notificacao="431490",
                dias_encerramento=10,
            ),
            self._notificacao(
                "poa",
                "431490",
                municipio_notificacao="431490",
                dias_encerramento=10,
            ),
            self._notificacao(
                "fora",
                "999999",
                dias_encerramento=10,
            ),
        )

        resultado = self.service.processar(
            self.municipios,
            notificacoes,
            date(2026, 1, 1),
            date(2026, 1, 31),
            date(2026, 1, 1),
            date(2026, 1, 31),
            ignorar_poa=True,
        )
        por_codigo = {
            item.codigo_ibge: item
            for item in resultado.municipios
        }

        acegua = por_codigo["430003"]
        self.assertEqual(acegua.total_notificados, 8)
        self.assertEqual(acegua.total_encerrados, 4)
        self.assertEqual(acegua.encerrados_no_prazo, 3)
        self.assertEqual(acegua.casos_nao_oportunos, 5)
        self.assertEqual(acegua.encerrados_apos_60_dias, 1)
        self.assertEqual(acegua.sem_data_encerramento, 1)
        self.assertEqual(acegua.classificacao_nao_valida, 2)
        self.assertEqual(acegua.inconclusivos_finais, 1)
        self.assertEqual(acegua.total_data_invalida, 1)
        self.assertEqual(acegua.casos_fora_prazo, 5)
        self.assertEqual(acegua.total_esquecidos, 1)
        self.assertEqual(acegua.percentual_oportunidade, 37.5)

        agua_santa = por_codigo["430005"]
        self.assertEqual(agua_santa.total_notificados, 0)
        self.assertEqual(
            agua_santa.percentual_oportunidade,
            "Sem Casos",
        )

        porto_alegre = por_codigo["431490"]
        self.assertEqual(porto_alegre.total_notificados, 1)
        self.assertEqual(porto_alegre.encerrados_no_prazo, 1)

        self.assertEqual(resultado.total_notificados, 9)
        self.assertEqual(resultado.total_encerrados, 5)
        self.assertEqual(resultado.total_no_prazo, 4)
        self.assertEqual(resultado.total_nao_oportunos, 5)
        self.assertEqual(resultado.percentual_estadual, 44.44)
        self.assertEqual(len(resultado.avisos), 3)

    def test_aplica_duplo_filtro_inclusivo(self):
        notificacoes = (
            self._notificacao(
                "dentro",
                "430003",
                notificacao="2026-01-10",
                sintomas="2025-12-20",
            ),
            self._notificacao(
                "notificacao-fora",
                "430003",
                notificacao="2026-02-01",
                sintomas="2025-12-20",
            ),
            self._notificacao(
                "sintomas-fora",
                "430003",
                notificacao="2026-01-10",
                sintomas="2026-01-01",
            ),
            self._notificacao(
                "limites",
                "430003",
                notificacao="2026-01-31",
                sintomas="2025-12-31",
            ),
        )

        resultado = self.service.processar(
            self.municipios,
            notificacoes,
            date(2026, 1, 1),
            date(2026, 1, 31),
            date(2025, 12, 1),
            date(2025, 12, 31),
        )

        self.assertEqual(resultado.total_notificados, 2)
        self.assertEqual(resultado.total_no_prazo, 2)

    def test_sentinela_exige_municipio_de_notificacao_no_dbf(self):
        with tempfile.TemporaryDirectory() as temporario:
            caminho = Path(temporario) / "DENGON26.dbf"
            caminho.write_bytes(b"dbf-sintetico")
            registro = {
                "NU_NOTIFIC": "1",
                "ID_MN_RESI": "430003",
                "DT_SIN_PRI": date(2026, 1, 10),
                "DT_NOTIFIC": date(2026, 1, 10),
                "DT_ENCERRA": date(2026, 1, 20),
                "CLASSI_FIN": 5,
            }

            notificacoes = self.service.carregar_notificacoes_dbf(
                [caminho],
                leitor_dbf=lambda _caminho: [registro],
            )
            self.assertEqual(len(notificacoes), 1)

            with self.assertRaisesRegex(ValueError, "ID_MUNICIP"):
                self.service.carregar_notificacoes_dbf(
                    [caminho],
                    exigir_municipio_notificacao=True,
                    leitor_dbf=lambda _caminho: [registro],
                )

    def test_normaliza_classificacoes_do_dbf(self):
        casos = {
            5: 5,
            10.0: 10,
            "11": 11,
            "12.0": 12,
            "8": 8,
            "ignorado": None,
            None: None,
        }
        for entrada, esperado in casos.items():
            with self.subTest(entrada=entrada):
                self.assertEqual(
                    self.service.normalizar_classificacao(entrada),
                    esperado,
                )

    def test_exporta_as_tres_abas_e_colunas_legadas(self):
        resultado = self.service.processar(
            self.municipios,
            (self._notificacao("1", "430003"),),
            date(2026, 1, 1),
            date(2026, 1, 31),
            date(2026, 1, 1),
            date(2026, 1, 31),
        )

        with tempfile.TemporaryDirectory() as temporario:
            destino = Path(temporario) / "relatorio_60d.xlsx"
            self.service.exportar_excel(resultado, destino)

            workbook = load_workbook(
                destino,
                read_only=True,
                data_only=True,
            )
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    ["Resumo_Estadual", "Dados_Municipios", "Legenda"],
                )
                dados = workbook["Dados_Municipios"]
                cabecalhos = tuple(
                    celula.value
                    for celula in next(dados.iter_rows(max_row=1))
                )
                self.assertEqual(
                    cabecalhos,
                    Relatorio60dService.CABECALHOS_SAIDA,
                )
                self.assertEqual(dados.max_row, 4)
                self.assertEqual(workbook["Resumo_Estadual"].max_row, 6)
                self.assertEqual(workbook["Legenda"].max_row, 16)
            finally:
                workbook.close()

    def test_preserva_relatorio_anterior_se_publicacao_falhar(self):
        resultado = self.service.processar(
            self.municipios,
            (),
            date(2026, 1, 1),
            date(2026, 1, 31),
            date(2026, 1, 1),
            date(2026, 1, 31),
        )

        with tempfile.TemporaryDirectory() as temporario:
            raiz = Path(temporario)
            destino = raiz / "relatorio_60d.xlsx"
            conteudo_anterior = b"relatorio-anterior"
            destino.write_bytes(conteudo_anterior)

            with patch(
                "app.services.qualifica.relatorio_60d_service.os.replace",
                side_effect=PermissionError("arquivo aberto"),
            ):
                with self.assertRaises(PermissionError):
                    self.service.exportar_excel(resultado, destino)

            self.assertEqual(destino.read_bytes(), conteudo_anterior)
            self.assertEqual(tuple(raiz.glob("*.tmp.xlsx")), ())

    def test_orquestra_arquivos_sem_expor_identificadores(self):
        with tempfile.TemporaryDirectory() as temporario:
            raiz = Path(temporario)
            dicionario = raiz / "municipios.xlsx"
            dbf = raiz / "DENGON26.dbf"
            saida = raiz / "resultado_60d.xlsx"
            dbf.write_bytes(b"dbf-sintetico")
            self._criar_dicionario(
                dicionario,
                [(430003, "Aceguá", 7, 4981)],
            )
            identificador = "identificador-ficticio"
            registro = {
                "NU_NOTIFIC": identificador,
                "ID_MN_RESI": "430003",
                "ID_MUNICIP": "430003",
                "DT_SIN_PRI": date(2026, 1, 10),
                "DT_NOTIFIC": date(2026, 1, 10),
                "DT_ENCERRA": date(2026, 2, 10),
                "CLASSI_FIN": 5,
            }
            mensagens: list[str] = []

            resultado = self.service.gerar_relatorio(
                caminho_dicionario=dicionario,
                caminhos_dbf=[dbf],
                notificacao_inicial=date(2026, 1, 1),
                notificacao_final=date(2026, 1, 31),
                sintomas_inicial=date(2026, 1, 1),
                sintomas_final=date(2026, 1, 31),
                caminho_saida=saida,
                callback_status=mensagens.append,
                leitor_dbf=lambda _caminho: [registro],
            )

            self.assertTrue(saida.is_file())
            self.assertEqual(resultado.total_notificados, 1)
            self.assertEqual(resultado.total_no_prazo, 1)
            self.assertEqual(len(mensagens), 5)
            self.assertFalse(
                any(
                    identificador in mensagem
                    for mensagem in mensagens
                )
            )

    @staticmethod
    def _notificacao(
        numero: str,
        residencia: str,
        *,
        municipio_notificacao: str = "430003",
        notificacao: str = "2026-01-10",
        sintomas: str = "2026-01-10",
        dias_encerramento: int | None = 10,
        classificacao: int = 5,
    ) -> Notificacao60d:
        data_notificacao = Relatorio60dService._normalizar_data_limite(
            notificacao,
            "notificação",
        )
        data_encerramento = (
            None
            if dias_encerramento is None
            else date.fromordinal(
                data_notificacao.toordinal() + dias_encerramento
            )
        )
        return Notificacao60d(
            numero=numero,
            municipio_residencia=residencia,
            municipio_notificacao=municipio_notificacao,
            data_primeiros_sintomas=(
                Relatorio60dService._normalizar_data_limite(
                    sintomas,
                    "sintomas",
                )
            ),
            data_notificacao=data_notificacao,
            data_encerramento=data_encerramento,
            classificacao_final=classificacao,
        )

    @staticmethod
    def _criar_dicionario(caminho: Path, linhas: list[tuple]):
        workbook = Workbook()
        planilha = workbook.active
        planilha.append(
            [
                "Cód Município",
                "Nome Município",
                "CRS",
                "Pop Município",
            ]
        )
        for linha in linhas:
            planilha.append(linha)
        workbook.save(caminho)
        workbook.close()


if __name__ == "__main__":
    unittest.main()
