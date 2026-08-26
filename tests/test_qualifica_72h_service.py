from __future__ import annotations

import tempfile
import unittest
from datetime import date
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from app.services.qualifica.relatorio_72h_service import (
    MunicipioQualifica,
    Notificacao72h,
    Relatorio72hService,
)


class Relatorio72hServiceTestCase(unittest.TestCase):
    def setUp(self):
        self.service = Relatorio72hService()
        self.municipios = (
            MunicipioQualifica("430003", "Aceguá", 7),
            MunicipioQualifica("430005", "Água Santa", 6),
            MunicipioQualifica("430010", "Agudo", 4),
        )

    def test_calcula_mesma_regra_de_zero_a_tres_dias(self):
        notificacoes = (
            self._notificacao(
                "1", "430003", "2026-01-10", "2026-01-10", "2026-01-13"
            ),
            self._notificacao(
                "2", "430003", "2026-01-10", "2026-01-10", "2026-01-14"
            ),
            self._notificacao(
                "3", "430003", "2026-01-10", "2026-01-10", "2026-01-09"
            ),
            self._notificacao(
                "4", "430003", "2026-01-10", "2026-01-10", None
            ),
            self._notificacao(
                "1", "430005", "2026-01-20", "2026-01-20", "2026-01-20"
            ),
            self._notificacao(
                "6", "430003", "2025-12-31", "2025-12-31", "2026-01-01"
            ),
            self._notificacao(
                "7", "999999", "2026-01-15", "2026-01-15", "2026-01-15"
            ),
        )

        resultado = self.service.processar(
            self.municipios,
            notificacoes,
            date(2026, 1, 4),
            date(2026, 1, 31),
        )
        por_codigo = {
            item.codigo_ibge: item
            for item in resultado.municipios
        }

        acegua = por_codigo["430003"]
        self.assertEqual(acegua.numero_notificacoes, 4)
        self.assertEqual(acegua.dias_dentro_do_prazo, 1)
        self.assertEqual(acegua.fora_prazo_72h, 3)
        self.assertEqual(acegua.percentual, 25.0)

        agua_santa = por_codigo["430005"]
        self.assertEqual(agua_santa.numero_notificacoes, 1)
        self.assertEqual(agua_santa.dias_dentro_do_prazo, 1)
        self.assertEqual(agua_santa.percentual, 100.0)

        agudo = por_codigo["430010"]
        self.assertEqual(agudo.numero_notificacoes, 0)
        self.assertEqual(agudo.percentual, 0.0)

        self.assertEqual(resultado.total_notificacoes, 5)
        self.assertEqual(resultado.total_dentro_do_prazo, 2)
        self.assertEqual(resultado.percentual_estadual, 40.0)
        self.assertEqual(len(resultado.avisos), 2)

    def test_carrega_dicionario_e_rejeita_ibge_truncado(self):
        with tempfile.TemporaryDirectory() as temporario:
            raiz = Path(temporario)
            valido = raiz / "municipios.xlsx"
            self._criar_dicionario(
                valido,
                [
                    (430003, "Aceguá", 7, 4981),
                    (430005, "Água Santa", 6, 3738),
                ],
            )

            municipios = self.service.carregar_municipios(valido)
            self.assertEqual(len(municipios), 2)
            self.assertEqual(municipios[0].codigo_ibge, "430003")

            invalido = raiz / "municipios_invalidos.xlsx"
            self._criar_dicionario(
                invalido,
                [(4300039, "Código inválido", 1, 1)],
            )
            with self.assertRaisesRegex(ValueError, "Código IBGE inválido"):
                self.service.carregar_municipios(invalido)

    def test_leitura_dbf_pode_ser_testada_sem_dado_real(self):
        with tempfile.TemporaryDirectory() as temporario:
            caminho = Path(temporario) / "DENGON26.dbf"
            caminho.write_bytes(b"dbf-sintetico")
            registros = [
                {
                    "NU_NOTIFIC": "123",
                    "ID_MUNICIP": "430003",
                    "DT_SIN_PRI": date(2026, 1, 10),
                    "DT_NOTIFIC": date(2026, 1, 10),
                    "DT_DIGITA": date(2026, 1, 12),
                }
            ]

            notificacoes = self.service.carregar_notificacoes_dbf(
                [caminho],
                leitor_dbf=lambda _caminho: registros,
            )

            self.assertEqual(len(notificacoes), 1)
            self.assertEqual(notificacoes[0].numero, "123")
            self.assertEqual(
                notificacoes[0].data_digitacao,
                date(2026, 1, 12),
            )

    def test_exportacao_preserva_arquivo_anterior_se_publicacao_falhar(self):
        resultado = self.service.processar(
            self.municipios,
            (),
            date(2026, 1, 4),
            date(2026, 1, 31),
        )

        with tempfile.TemporaryDirectory() as temporario:
            raiz = Path(temporario)
            destino = raiz / "relatorio.xlsx"
            conteudo_anterior = b"arquivo-anterior-preservado"
            destino.write_bytes(conteudo_anterior)

            with patch(
                "app.services.qualifica.relatorio_72h_service.os.replace",
                side_effect=PermissionError("arquivo aberto"),
            ):
                with self.assertRaises(PermissionError):
                    self.service.exportar_excel(resultado, destino)

            self.assertEqual(destino.read_bytes(), conteudo_anterior)
            self.assertEqual(tuple(raiz.glob("*.tmp.xlsx")), ())

    def test_exporta_abas_e_colunas_compativeis_com_programa_anterior(self):
        resultado = self.service.processar(
            self.municipios,
            (
                self._notificacao(
                    "1",
                    "430003",
                    "2026-01-10",
                    "2026-01-10",
                    "2026-01-11",
                ),
            ),
            date(2026, 1, 4),
            date(2026, 1, 31),
        )

        with tempfile.TemporaryDirectory() as temporario:
            destino = Path(temporario) / "relatorio.xlsx"
            self.service.exportar_excel(resultado, destino)

            workbook = load_workbook(
                destino,
                read_only=True,
                data_only=True,
            )
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    ["Resumo_Estadual", "Dados_Municipios"],
                )
                dados = workbook["Dados_Municipios"]
                cabecalhos = tuple(
                    celula.value
                    for celula in next(dados.iter_rows(max_row=1))
                )
                self.assertEqual(
                    cabecalhos,
                    Relatorio72hService.CABECALHOS_SAIDA,
                )
                self.assertEqual(dados.max_row, 4)
            finally:
                workbook.close()

    def test_orquestra_arquivos_e_emite_status_sem_dados_individuais(self):
        with tempfile.TemporaryDirectory() as temporario:
            raiz = Path(temporario)
            dicionario = raiz / "municipios.xlsx"
            dbf = raiz / "DENGON26.dbf"
            saida = raiz / "resultado.xlsx"
            dbf.write_bytes(b"dbf-sintetico")
            self._criar_dicionario(
                dicionario,
                [(430003, "Aceguá", 7, 4981)],
            )
            registros = [
                {
                    "NU_NOTIFIC": "identificador-ficticio",
                    "ID_MUNICIP": "430003",
                    "DT_SIN_PRI": date(2026, 1, 10),
                    "DT_NOTIFIC": date(2026, 1, 10),
                    "DT_DIGITA": date(2026, 1, 12),
                }
            ]
            mensagens: list[str] = []

            resultado = self.service.gerar_relatorio(
                caminho_dicionario=dicionario,
                caminhos_dbf=[dbf],
                data_inicial=date(2026, 1, 4),
                data_final=date(2026, 1, 31),
                caminho_saida=saida,
                callback_status=mensagens.append,
                leitor_dbf=lambda _caminho: registros,
            )

            self.assertTrue(saida.is_file())
            self.assertEqual(resultado.total_notificacoes, 1)
            self.assertEqual(len(mensagens), 5)
            self.assertFalse(
                any(
                    "identificador-ficticio" in mensagem
                    for mensagem in mensagens
                )
            )

    @staticmethod
    def _notificacao(
        numero: str,
        municipio: str,
        sintomas: str,
        notificacao: str,
        digitacao: str | None,
    ) -> Notificacao72h:
        converter = Relatorio72hService.converter_data
        return Notificacao72h(
            numero=numero,
            municipio_notificacao=municipio,
            data_primeiros_sintomas=converter(sintomas),
            data_notificacao=converter(notificacao),
            data_digitacao=converter(digitacao),
        )

    @staticmethod
    def _criar_dicionario(caminho: Path, linhas: list[tuple]):
        workbook = Workbook()
        planilha = workbook.active
        planilha.append(
            [
                "Cód Município",
                "Nome Município",
                "CRS",
                "Pop Município",
            ]
        )
        for linha in linhas:
            planilha.append(linha)
        workbook.save(caminho)
        workbook.close()


if __name__ == "__main__":
    unittest.main()
