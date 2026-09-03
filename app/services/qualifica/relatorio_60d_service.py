from __future__ import annotations

import math
import os
import re
import tempfile
import unicodedata
from collections import Counter
from collections.abc import Callable, Iterable, Mapping, Sequence
from dataclasses import dataclass
from datetime import date
from itertools import chain
from pathlib import Path
from typing import Any, Iterator

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.services.qualifica.relatorio_72h_service import (
    MunicipioQualifica,
    Relatorio72hService,
)


CLASSIFICACOES_VALIDAS = frozenset({5, 10, 11, 12})
CODIGO_IBGE_PORTO_ALEGRE = "431490"
COLUNAS_DBF_OBRIGATORIAS = frozenset(
    {
        "ID_MN_RESI",
        "DT_SIN_PRI",
        "DT_NOTIFIC",
        "DT_ENCERRA",
        "CLASSI_FIN",
    }
)


@dataclass(frozen=True)
class Notificacao60d:
    """Campos mínimos de uma notificação usados no indicador de 60 dias."""

    numero: str
    municipio_residencia: str
    municipio_notificacao: str
    data_primeiros_sintomas: date | None
    data_notificacao: date | None
    data_encerramento: date | None
    classificacao_final: int | None


@dataclass(frozen=True)
class ResultadoMunicipio60d:
    """Resultado agregado do indicador para um município."""

    codigo_ibge: str
    nome: str
    crs: int
    total_notificados: int
    total_encerrados: int
    encerrados_no_prazo: int
    casos_nao_oportunos: int
    encerrados_apos_60_dias: int
    sem_data_encerramento: int
    classificacao_nao_valida: int
    inconclusivos_finais: int
    total_data_invalida: int
    casos_fora_prazo: int
    total_esquecidos: int
    percentual_oportunidade: float | str


@dataclass(frozen=True)
class ResultadoRelatorio60d:
    """Resultado completo, contendo exclusivamente dados agregados."""

    notificacao_inicial: date
    notificacao_final: date
    sintomas_inicial: date
    sintomas_final: date
    municipios: tuple[ResultadoMunicipio60d, ...]
    total_notificados: int
    total_encerrados: int
    total_no_prazo: int
    total_nao_oportunos: int
    percentual_estadual: float
    avisos: tuple[str, ...] = ()


@dataclass
class _ContagemMunicipio60d:
    total_notificados: int = 0
    total_encerrados: int = 0
    encerrados_no_prazo: int = 0
    encerrados_apos_60_dias: int = 0
    sem_data_encerramento: int = 0
    classificacao_nao_valida: int = 0
    inconclusivos_finais: int = 0
    total_data_invalida: int = 0


class Relatorio60dService:
    """
    Calcula a oportunidade de encerramento da Dengue em até 60 dias.

    O resultado mantém apenas agregados municipais. Os registros individuais
    são consumidos progressivamente e não são persistidos ou enviados à
    interface.
    """

    CABECALHOS_SAIDA = (
        "ID_MN_RESI",
        "Nome_Municipio",
        "CRS",
        "Total_Notificados",
        "Total_Encerrados",
        "Encerrados_No_Prazo",
        "Casos_Nao_Oportunos",
        "Encerrados_Apos_60_Dias",
        "Sem_Data_Encerramento",
        "Classificacao_Nao_Valida",
        "Inconclusivos_Finais",
        "Total_Data_Inv",
        "Casos_Fora_Prazo",
        "Total_Esquecidos",
        "Percentual_Oportunidade",
    )

    DESCRICOES_COLUNAS = {
        "ID_MN_RESI": (
            "Código IBGE de 6 dígitos do município de residência."
        ),
        "Nome_Municipio": "Nome padronizado do município.",
        "CRS": "Coordenadoria Regional de Saúde.",
        "Total_Notificados": "Total de casos incluídos na coorte.",
        "Total_Encerrados": (
            "Casos com data não negativa e classificação final válida."
        ),
        "Encerrados_No_Prazo": (
            "Casos encerrados validamente entre 0 e 60 dias após a "
            "notificação."
        ),
        "Casos_Nao_Oportunos": (
            "Complemento do numerador: Total_Notificados - "
            "Encerrados_No_Prazo."
        ),
        "Casos_Fora_Prazo": (
            "Nome mantido por compatibilidade; equivale a "
            "Casos_Nao_Oportunos."
        ),
        "Encerrados_Apos_60_Dias": (
            "Casos válidos encerrados depois de 60 dias."
        ),
        "Sem_Data_Encerramento": "Casos sem data de encerramento.",
        "Total_Esquecidos": (
            "Nome antigo mantido por compatibilidade; equivale a "
            "Sem_Data_Encerramento."
        ),
        "Classificacao_Nao_Valida": (
            "Casos com data preenchida, mas classificação não aceita no "
            "indicador."
        ),
        "Inconclusivos_Finais": "Casos com CLASSI_FIN = 8.",
        "Total_Data_Inv": (
            "Casos com encerramento anterior à notificação."
        ),
        "Percentual_Oportunidade": (
            "Encerrados_No_Prazo / Total_Notificados × 100."
        ),
    }

    def __init__(self):
        self._servico_comum = Relatorio72hService()

    def carregar_municipios(
        self,
        caminho_planilha: str | Path,
    ) -> tuple[MunicipioQualifica, ...]:
        """Reutiliza a validação institucional do dicionário do Qualifica."""

        return self._servico_comum.carregar_municipios(caminho_planilha)

    def carregar_notificacoes_dbf(
        self,
        caminhos_dbf: Sequence[str | Path],
        *,
        exigir_municipio_notificacao: bool = False,
        leitor_dbf: Callable[
            [Path],
            Iterable[Mapping[str, Any]],
        ]
        | None = None,
    ) -> tuple[Notificacao60d, ...]:
        """Materializa registros apenas quando a chamada exigir a coleção."""

        return tuple(
            self.iterar_notificacoes_dbf(
                caminhos_dbf,
                exigir_municipio_notificacao=(
                    exigir_municipio_notificacao
                ),
                leitor_dbf=leitor_dbf,
            )
        )

    def iterar_notificacoes_dbf(
        self,
        caminhos_dbf: Sequence[str | Path],
        *,
        exigir_municipio_notificacao: bool = False,
        leitor_dbf: Callable[
            [Path],
            Iterable[Mapping[str, Any]],
        ]
        | None = None,
    ) -> Iterator[Notificacao60d]:
        """Lê progressivamente somente os campos usados no indicador."""

        if not caminhos_dbf:
            raise ValueError(
                "Selecione pelo menos um banco DBF do SINAN."
            )

        leitor = leitor_dbf or self._abrir_dbf
        registros_encontrados = 0

        for caminho_bruto in caminhos_dbf:
            caminho = Path(caminho_bruto).expanduser()
            if not caminho.is_file():
                raise FileNotFoundError(
                    f"O banco DBF não foi encontrado: {caminho}"
                )
            if caminho.suffix.casefold() != ".dbf":
                raise ValueError(
                    f"O arquivo não é um DBF: {caminho.name}"
                )
            if caminho.stat().st_size <= 0:
                raise ValueError(
                    f"O banco DBF está vazio: {caminho.name}"
                )

            registros = iter(leitor(caminho))
            try:
                primeiro = next(registros)
            except StopIteration:
                continue

            self._validar_colunas_dbf(
                primeiro,
                exigir_municipio_notificacao=(
                    exigir_municipio_notificacao
                ),
            )
            for registro in chain((primeiro,), registros):
                registros_encontrados += 1
                yield self._converter_registro_dbf(registro)

        if registros_encontrados == 0:
            raise ValueError(
                "Nenhum registro foi encontrado nos bancos DBF."
            )

    def processar(
        self,
        municipios: Sequence[MunicipioQualifica],
        notificacoes: Iterable[Notificacao60d],
        notificacao_inicial: date,
        notificacao_final: date,
        sintomas_inicial: date,
        sintomas_final: date,
        *,
        ignorar_poa: bool = False,
    ) -> ResultadoRelatorio60d:
        notificacao_inicial = self._normalizar_data_limite(
            notificacao_inicial,
            "inicial de notificação",
        )
        notificacao_final = self._normalizar_data_limite(
            notificacao_final,
            "final de notificação",
        )
        sintomas_inicial = self._normalizar_data_limite(
            sintomas_inicial,
            "inicial de sintomas",
        )
        sintomas_final = self._normalizar_data_limite(
            sintomas_final,
            "final de sintomas",
        )
        if notificacao_inicial > notificacao_final:
            raise ValueError(
                "A data inicial de notificação é posterior à data final."
            )
        if sintomas_inicial > sintomas_final:
            raise ValueError(
                "A data inicial de sintomas é posterior à data final."
            )

        municipios_por_codigo = {
            municipio.codigo_ibge: municipio
            for municipio in municipios
        }
        if len(municipios_por_codigo) != len(municipios):
            raise ValueError(
                "A lista de municípios possui códigos IBGE duplicados."
            )

        contagens = {
            codigo: _ContagemMunicipio60d()
            for codigo in municipios_por_codigo
        }
        numeros_notificacao: Counter[str] = Counter()
        fora_do_dicionario = 0
        excluidas_sentinela = 0

        for notificacao in notificacoes:
            residencia = Relatorio72hService.normalizar_codigo_ibge(
                notificacao.municipio_residencia,
                permitir_vazio=True,
            )
            notificante = Relatorio72hService.normalizar_codigo_ibge(
                notificacao.municipio_notificacao,
                permitir_vazio=True,
            )
            if (
                ignorar_poa
                and notificante == CODIGO_IBGE_PORTO_ALEGRE
                and residencia != CODIGO_IBGE_PORTO_ALEGRE
            ):
                excluidas_sentinela += 1
                continue

            data_notificacao = notificacao.data_notificacao
            if (
                data_notificacao is None
                or data_notificacao < notificacao_inicial
                or data_notificacao > notificacao_final
            ):
                continue

            data_sintomas = notificacao.data_primeiros_sintomas
            if (
                data_sintomas is None
                or data_sintomas < sintomas_inicial
                or data_sintomas > sintomas_final
            ):
                continue

            if residencia not in contagens:
                fora_do_dicionario += 1
                continue

            contagem = contagens[residencia]
            contagem.total_notificados += 1

            numero = Relatorio72hService.normalizar_numero_notificacao(
                notificacao.numero
            )
            if numero:
                numeros_notificacao[numero] += 1

            classificacao = notificacao.classificacao_final
            if classificacao == 8:
                contagem.inconclusivos_finais += 1

            data_encerramento = notificacao.data_encerramento
            sem_data = data_encerramento is None
            dias_encerramento = (
                None
                if sem_data
                else (data_encerramento - data_notificacao).days
            )
            data_invalida = (
                dias_encerramento is not None
                and dias_encerramento < 0
            )
            classificacao_valida = (
                classificacao in CLASSIFICACOES_VALIDAS
            )
            encerrado_valido = (
                not sem_data
                and not data_invalida
                and classificacao_valida
            )
            no_prazo = (
                encerrado_valido
                and dias_encerramento is not None
                and 0 <= dias_encerramento <= 60
            )
            encerrado_apos_prazo = (
                encerrado_valido
                and dias_encerramento is not None
                and dias_encerramento > 60
            )
            classificacao_nao_valida = (
                not sem_data
                and not data_invalida
                and not classificacao_valida
            )

            if encerrado_valido:
                contagem.total_encerrados += 1
            if no_prazo:
                contagem.encerrados_no_prazo += 1
            elif sem_data:
                contagem.sem_data_encerramento += 1
            elif data_invalida:
                contagem.total_data_invalida += 1
            elif classificacao_nao_valida:
                contagem.classificacao_nao_valida += 1
            elif encerrado_apos_prazo:
                contagem.encerrados_apos_60_dias += 1
            else:
                raise ValueError(
                    "Não foi possível classificar um caso não oportuno."
                )

        resultados: list[ResultadoMunicipio60d] = []
        for codigo, municipio in municipios_por_codigo.items():
            contagem = contagens[codigo]
            casos_nao_oportunos = (
                contagem.total_notificados
                - contagem.encerrados_no_prazo
            )
            percentual: float | str = (
                round(
                    (
                        contagem.encerrados_no_prazo
                        / contagem.total_notificados
                    )
                    * 100,
                    2,
                )
                if contagem.total_notificados > 0
                else "Sem Casos"
            )
            resultados.append(
                ResultadoMunicipio60d(
                    codigo_ibge=codigo,
                    nome=municipio.nome,
                    crs=municipio.crs,
                    total_notificados=contagem.total_notificados,
                    total_encerrados=contagem.total_encerrados,
                    encerrados_no_prazo=(
                        contagem.encerrados_no_prazo
                    ),
                    casos_nao_oportunos=casos_nao_oportunos,
                    encerrados_apos_60_dias=(
                        contagem.encerrados_apos_60_dias
                    ),
                    sem_data_encerramento=(
                        contagem.sem_data_encerramento
                    ),
                    classificacao_nao_valida=(
                        contagem.classificacao_nao_valida
                    ),
                    inconclusivos_finais=(
                        contagem.inconclusivos_finais
                    ),
                    total_data_invalida=(
                        contagem.total_data_invalida
                    ),
                    casos_fora_prazo=casos_nao_oportunos,
                    total_esquecidos=contagem.sem_data_encerramento,
                    percentual_oportunidade=percentual,
                )
            )

        resultados.sort(
            key=lambda item: (
                item.crs,
                self._texto_para_ordenacao(item.nome),
            )
        )
        total_notificados = sum(
            item.total_notificados for item in resultados
        )
        total_encerrados = sum(
            item.total_encerrados for item in resultados
        )
        total_no_prazo = sum(
            item.encerrados_no_prazo for item in resultados
        )
        total_nao_oportunos = total_notificados - total_no_prazo
        percentual_estadual = (
            round((total_no_prazo / total_notificados) * 100, 2)
            if total_notificados > 0
            else 0.0
        )

        avisos: list[str] = []
        repetidos = sum(
            1
            for quantidade in numeros_notificacao.values()
            if quantidade > 1
        )
        if repetidos:
            avisos.append(
                f"Foram encontrados {repetidos} número(s) de "
                "notificação repetido(s). As linhas foram mantidas."
            )
        if fora_do_dicionario:
            avisos.append(
                f"Foram desconsideradas {fora_do_dicionario} "
                "notificação(ões) com município de residência ausente "
                "ou fora do dicionário."
            )
        if excluidas_sentinela:
            avisos.append(
                f"A regra Sentinela excluiu {excluidas_sentinela} "
                "notificação(ões) de Porto Alegre para residentes de "
                "outros municípios."
            )

        return ResultadoRelatorio60d(
            notificacao_inicial=notificacao_inicial,
            notificacao_final=notificacao_final,
            sintomas_inicial=sintomas_inicial,
            sintomas_final=sintomas_final,
            municipios=tuple(resultados),
            total_notificados=total_notificados,
            total_encerrados=total_encerrados,
            total_no_prazo=total_no_prazo,
            total_nao_oportunos=total_nao_oportunos,
            percentual_estadual=percentual_estadual,
            avisos=tuple(avisos),
        )

    def exportar_excel(
        self,
        resultado: ResultadoRelatorio60d,
        caminho_saida: str | Path,
    ) -> Path:
        destino = Path(caminho_saida).expanduser()
        if destino.suffix.casefold() != ".xlsx":
            raise ValueError(
                "O relatório de 60 dias precisa ser salvo como XLSX."
            )
        destino.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        resumo = workbook.active
        resumo.title = "Resumo_Estadual"
        dados = workbook.create_sheet("Dados_Municipios")
        legenda = workbook.create_sheet("Legenda")

        resumo.append(["Métrica (Estado do RS)", "Valor"])
        resumo.append(
            [
                "Total de Notificados (Coorte por Sintomas)",
                resultado.total_notificados,
            ]
        )
        resumo.append(
            ["Total de Encerrados Válidos", resultado.total_encerrados]
        )
        resumo.append(
            [
                "Encerrados Oportunamente (Até 60 dias)",
                resultado.total_no_prazo,
            ]
        )
        resumo.append(
            ["Casos Não Oportunos", resultado.total_nao_oportunos]
        )
        resumo.append(
            [
                "Percentual de Oportunidade Estadual (%)",
                resultado.percentual_estadual,
            ]
        )

        dados.append(list(self.CABECALHOS_SAIDA))
        for municipio in resultado.municipios:
            dados.append(
                [
                    municipio.codigo_ibge,
                    municipio.nome,
                    municipio.crs,
                    municipio.total_notificados,
                    municipio.total_encerrados,
                    municipio.encerrados_no_prazo,
                    municipio.casos_nao_oportunos,
                    municipio.encerrados_apos_60_dias,
                    municipio.sem_data_encerramento,
                    municipio.classificacao_nao_valida,
                    municipio.inconclusivos_finais,
                    municipio.total_data_invalida,
                    municipio.casos_fora_prazo,
                    municipio.total_esquecidos,
                    municipio.percentual_oportunidade,
                ]
            )

        legenda.append(["Nome da Coluna", "Descrição"])
        for nome, descricao in self.DESCRICOES_COLUNAS.items():
            legenda.append([nome, descricao])

        self._formatar_planilha_resumo(resumo)
        self._formatar_planilha_municipios(dados)
        self._formatar_planilha_legenda(legenda)
        try:
            self._salvar_workbook_atomico(workbook, destino)
        finally:
            workbook.close()
        return destino

    def gerar_relatorio(
        self,
        caminho_dicionario: str | Path,
        caminhos_dbf: Sequence[str | Path],
        notificacao_inicial: date,
        notificacao_final: date,
        sintomas_inicial: date,
        sintomas_final: date,
        caminho_saida: str | Path,
        *,
        ignorar_poa: bool = False,
        callback_status: Callable[[str], None] | None = None,
        leitor_dbf: Callable[
            [Path],
            Iterable[Mapping[str, Any]],
        ]
        | None = None,
    ) -> ResultadoRelatorio60d:
        destino = Path(caminho_saida).expanduser()
        origens = {
            Path(caminho_dicionario).expanduser().resolve(),
            *(
                Path(caminho).expanduser().resolve()
                for caminho in caminhos_dbf
            ),
        }
        if destino.resolve() in origens:
            raise ValueError(
                "O arquivo de saída não pode substituir um arquivo de "
                "entrada."
            )

        self._emitir_status(
            callback_status,
            "Validando o dicionário de municípios.",
        )
        municipios = self.carregar_municipios(caminho_dicionario)

        self._emitir_status(
            callback_status,
            f"Lendo {len(caminhos_dbf)} banco(s) DBF do SINAN.",
        )
        notificacoes = self.iterar_notificacoes_dbf(
            caminhos_dbf,
            exigir_municipio_notificacao=ignorar_poa,
            leitor_dbf=leitor_dbf,
        )

        self._emitir_status(
            callback_status,
            "Calculando a oportunidade de encerramento em 60 dias.",
        )
        resultado = self.processar(
            municipios=municipios,
            notificacoes=notificacoes,
            notificacao_inicial=notificacao_inicial,
            notificacao_final=notificacao_final,
            sintomas_inicial=sintomas_inicial,
            sintomas_final=sintomas_final,
            ignorar_poa=ignorar_poa,
        )

        self._emitir_status(
            callback_status,
            "Gerando o relatório consolidado em Excel.",
        )
        self.exportar_excel(resultado, destino)
        self._emitir_status(
            callback_status,
            "Relatório de 60 dias concluído.",
        )
        return resultado

    @staticmethod
    def normalizar_classificacao(valor: Any) -> int | None:
        if valor in (None, ""):
            return None
        if isinstance(valor, (int, float)):
            numero = float(valor)
            if math.isfinite(numero) and numero.is_integer():
                return int(numero)
            return None

        texto = str(valor).strip()
        if not re.fullmatch(r"[+-]?\d+(?:\.0+)?", texto):
            return None
        try:
            numero = float(texto)
        except ValueError:
            return None
        return int(numero) if numero.is_integer() else None

    @classmethod
    def _converter_registro_dbf(
        cls,
        registro: Mapping[str, Any],
    ) -> Notificacao60d:
        return Notificacao60d(
            numero=Relatorio72hService.normalizar_numero_notificacao(
                registro.get("NU_NOTIFIC")
            ),
            municipio_residencia=(
                Relatorio72hService.normalizar_codigo_ibge(
                    registro.get("ID_MN_RESI"),
                    permitir_vazio=True,
                )
            ),
            municipio_notificacao=(
                Relatorio72hService.normalizar_codigo_ibge(
                    registro.get("ID_MUNICIP"),
                    permitir_vazio=True,
                )
            ),
            data_primeiros_sintomas=Relatorio72hService.converter_data(
                registro.get("DT_SIN_PRI")
            ),
            data_notificacao=Relatorio72hService.converter_data(
                registro.get("DT_NOTIFIC")
            ),
            data_encerramento=Relatorio72hService.converter_data(
                registro.get("DT_ENCERRA")
            ),
            classificacao_final=cls.normalizar_classificacao(
                registro.get("CLASSI_FIN")
            ),
        )

    @staticmethod
    def _abrir_dbf(
        caminho: Path,
    ) -> Iterable[Mapping[str, Any]]:
        try:
            from dbfread import DBF
        except ImportError as erro:
            raise RuntimeError(
                "A dependência dbfread não está instalada. "
                "Reinstale as dependências do ArboHub."
            ) from erro

        return DBF(
            str(caminho),
            encoding="latin1",
            load=False,
            char_decode_errors="strict",
        )

    @staticmethod
    def _validar_colunas_dbf(
        registro: Mapping[str, Any],
        *,
        exigir_municipio_notificacao: bool,
    ):
        obrigatorias = set(COLUNAS_DBF_OBRIGATORIAS)
        if exigir_municipio_notificacao:
            obrigatorias.add("ID_MUNICIP")
        faltantes = sorted(obrigatorias - set(registro))
        if faltantes:
            raise ValueError(
                "O banco SINAN não contém as colunas necessárias: "
                + ", ".join(faltantes)
            )

    @staticmethod
    def _normalizar_data_limite(
        valor: date,
        descricao: str,
    ) -> date:
        data_convertida = Relatorio72hService.converter_data(valor)
        if data_convertida is None:
            raise ValueError(
                f"A data {descricao} do relatório é inválida."
            )
        return data_convertida

    @staticmethod
    def _texto_para_ordenacao(valor: str) -> str:
        texto = unicodedata.normalize("NFKD", valor.casefold())
        return "".join(
            caractere
            for caractere in texto
            if not unicodedata.combining(caractere)
        )

    @staticmethod
    def _emitir_status(
        callback: Callable[[str], None] | None,
        mensagem: str,
    ):
        if callback:
            callback(mensagem)

    @staticmethod
    def _formatar_cabecalho(planilha, intervalo: str):
        preenchimento = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )
        for celula in planilha[intervalo][0]:
            celula.fill = preenchimento
            celula.font = Font(color="FFFFFF", bold=True)
            celula.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

    @classmethod
    def _formatar_planilha_resumo(cls, planilha):
        cls._formatar_cabecalho(planilha, "A1:B1")
        planilha.column_dimensions["A"].width = 55
        planilha.column_dimensions["B"].width = 18
        planilha["B6"].number_format = "0.00"
        planilha.freeze_panes = "A2"
        planilha.auto_filter.ref = "A1:B6"

    @classmethod
    def _formatar_planilha_municipios(cls, planilha):
        cls._formatar_cabecalho(planilha, "A1:O1")
        larguras = {
            "A": 14,
            "B": 34,
            "C": 8,
            "D": 20,
            "E": 19,
            "F": 23,
            "G": 21,
            "H": 25,
            "I": 24,
            "J": 26,
            "K": 21,
            "L": 18,
            "M": 20,
            "N": 18,
            "O": 24,
        }
        for coluna, largura in larguras.items():
            planilha.column_dimensions[coluna].width = largura
        for celula in planilha["A"][1:]:
            celula.number_format = "@"
        for celula in planilha["O"][1:]:
            if isinstance(celula.value, (int, float)):
                celula.number_format = "0.00"
        planilha.freeze_panes = "A2"
        planilha.auto_filter.ref = planilha.dimensions

    @classmethod
    def _formatar_planilha_legenda(cls, planilha):
        cls._formatar_cabecalho(planilha, "A1:B1")
        planilha.column_dimensions["A"].width = 30
        planilha.column_dimensions["B"].width = 90
        planilha.freeze_panes = "A2"
        planilha.auto_filter.ref = planilha.dimensions

    @classmethod
    def _salvar_workbook_atomico(
        cls,
        workbook: Workbook,
        destino: Path,
    ):
        descritor, nome_temporario = tempfile.mkstemp(
            prefix=f".{destino.stem}_",
            suffix=".tmp.xlsx",
            dir=destino.parent,
        )
        os.close(descritor)
        temporario = Path(nome_temporario)

        try:
            workbook.save(temporario)
            verificacao = load_workbook(
                temporario,
                read_only=True,
                data_only=True,
            )
            try:
                abas_esperadas = [
                    "Resumo_Estadual",
                    "Dados_Municipios",
                    "Legenda",
                ]
                if verificacao.sheetnames != abas_esperadas:
                    raise ValueError(
                        "O relatório temporário não contém as abas "
                        "esperadas."
                    )
                dados = verificacao["Dados_Municipios"]
                cabecalhos = tuple(
                    celula.value
                    for celula in next(
                        dados.iter_rows(max_row=1)
                    )
                )
                if cabecalhos != cls.CABECALHOS_SAIDA:
                    raise ValueError(
                        "O relatório temporário não contém as colunas "
                        "esperadas."
                    )
            finally:
                verificacao.close()
            os.replace(temporario, destino)
        finally:
            temporario.unlink(missing_ok=True)
