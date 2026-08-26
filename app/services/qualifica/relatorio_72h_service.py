from __future__ import annotations

import math
import os
import re
import tempfile
import unicodedata
from collections import Counter
from collections.abc import Callable, Iterable, Mapping, Sequence
from dataclasses import dataclass
from datetime import date, datetime
from itertools import chain
from pathlib import Path
from typing import Any, Iterator

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


COLUNAS_DBF_OBRIGATORIAS = frozenset(
    {
        "NU_NOTIFIC",
        "ID_MUNICIP",
        "DT_SIN_PRI",
        "DT_NOTIFIC",
        "DT_DIGITA",
    }
)


@dataclass(frozen=True)
class MunicipioQualifica:
    """Município necessário para consolidação do indicador."""

    codigo_ibge: str
    nome: str
    crs: int


@dataclass(frozen=True)
class Notificacao72h:
    """Campos mínimos de uma notificação usados no cálculo."""

    numero: str
    municipio_notificacao: str
    data_primeiros_sintomas: date | None
    data_notificacao: date | None
    data_digitacao: date | None


@dataclass(frozen=True)
class ResultadoMunicipio72h:
    """Resultado agregado de um município, sem dados individuais."""

    codigo_ibge: str
    nome: str
    crs: int
    numero_notificacoes: int
    dias_dentro_do_prazo: int
    fora_prazo_72h: int
    percentual: float


@dataclass(frozen=True)
class ResultadoRelatorio72h:
    """Resultado completo que será apresentado e exportado."""

    data_inicial: date
    data_final: date
    municipios: tuple[ResultadoMunicipio72h, ...]
    total_notificacoes: int
    total_dentro_do_prazo: int
    percentual_estadual: float
    avisos: tuple[str, ...] = ()


class Relatorio72hService:
    """
    Calcula a oportunidade de digitação do SINAN em até três dias.

    O serviço mantém somente agregados municipais no resultado. Registros
    individuais são lidos durante o processamento e não são persistidos,
    registrados em log ou enviados para a interface.
    """

    CABECALHOS_DICIONARIO = {
        "cod municipio": "codigo_ibge",
        "nome municipio": "nome",
        "crs": "crs",
    }

    CABECALHOS_SAIDA = (
        "ID_MUNICIP",
        "Nome_Municipio",
        "CRS",
        "Numero_Notificacoes",
        "Dias_dentro_do_prazo",
        "Fora_Prazo_72h",
        "Percentual",
    )

    def carregar_municipios(
        self,
        caminho_planilha: str | Path,
    ) -> tuple[MunicipioQualifica, ...]:
        caminho = Path(caminho_planilha).expanduser()
        if not caminho.is_file():
            raise FileNotFoundError(
                f"O dicionário de municípios não foi encontrado: {caminho}"
            )
        if caminho.suffix.casefold() != ".xlsx":
            raise ValueError(
                "O dicionário de municípios precisa ser um arquivo XLSX."
            )

        workbook = load_workbook(
            caminho,
            read_only=True,
            data_only=True,
        )
        try:
            planilha = workbook.active
            linhas = planilha.iter_rows(values_only=True)
            try:
                cabecalhos_brutos = next(linhas)
            except StopIteration as erro:
                raise ValueError(
                    "O dicionário de municípios está vazio."
                ) from erro

            indices: dict[str, int] = {}
            for indice, valor in enumerate(cabecalhos_brutos):
                normalizado = self._normalizar_cabecalho(valor)
                destino = self.CABECALHOS_DICIONARIO.get(normalizado)
                if destino:
                    indices[destino] = indice

            faltantes = sorted(
                set(self.CABECALHOS_DICIONARIO.values())
                - set(indices)
            )
            if faltantes:
                raise ValueError(
                    "O dicionário não contém as colunas necessárias: "
                    + ", ".join(faltantes)
                )

            municipios: list[MunicipioQualifica] = []
            codigos: set[str] = set()

            for numero_linha, linha in enumerate(linhas, start=2):
                if not any(
                    valor not in (None, "")
                    for valor in linha
                ):
                    continue

                codigo = self.normalizar_codigo_ibge(
                    self._valor_linha(
                        linha,
                        indices["codigo_ibge"],
                    )
                )
                if not codigo:
                    raise ValueError(
                        "Código IBGE ausente no dicionário, "
                        f"linha {numero_linha}."
                    )
                if codigo in codigos:
                    raise ValueError(
                        "O dicionário possui código IBGE duplicado: "
                        f"{codigo}."
                    )

                nome = str(
                    self._valor_linha(
                        linha,
                        indices["nome"],
                    )
                    or ""
                ).strip()
                if not nome:
                    raise ValueError(
                        "Nome de município ausente no dicionário, "
                        f"linha {numero_linha}."
                    )

                crs = self._normalizar_crs(
                    self._valor_linha(
                        linha,
                        indices["crs"],
                    ),
                    numero_linha,
                )
                codigos.add(codigo)
                municipios.append(
                    MunicipioQualifica(
                        codigo_ibge=codigo,
                        nome=nome,
                        crs=crs,
                    )
                )
        finally:
            workbook.close()

        if not municipios:
            raise ValueError(
                "O dicionário não contém municípios válidos."
            )

        return tuple(municipios)

    def carregar_notificacoes_dbf(
        self,
        caminhos_dbf: Sequence[str | Path],
        *,
        leitor_dbf: Callable[
            [Path],
            Iterable[Mapping[str, Any]],
        ] | None = None,
    ) -> tuple[Notificacao72h, ...]:
        """Materializa registros quando uma chamada precisar da coleção."""

        return tuple(
            self.iterar_notificacoes_dbf(
                caminhos_dbf,
                leitor_dbf=leitor_dbf,
            )
        )

    def iterar_notificacoes_dbf(
        self,
        caminhos_dbf: Sequence[str | Path],
        *,
        leitor_dbf: Callable[
            [Path],
            Iterable[Mapping[str, Any]],
        ] | None = None,
    ) -> Iterator[Notificacao72h]:
        """Lê os registros progressivamente para limitar o uso de memória."""

        if not caminhos_dbf:
            raise ValueError(
                "Selecione pelo menos um banco DBF do SINAN."
            )

        leitor = leitor_dbf or self._abrir_dbf
        registros_encontrados = 0

        for caminho_bruto in caminhos_dbf:
            caminho = Path(caminho_bruto).expanduser()
            if not caminho.is_file():
                raise FileNotFoundError(
                    f"O banco DBF não foi encontrado: {caminho}"
                )
            if caminho.suffix.casefold() != ".dbf":
                raise ValueError(
                    f"O arquivo não é um DBF: {caminho.name}"
                )
            if caminho.stat().st_size <= 0:
                raise ValueError(
                    f"O banco DBF está vazio: {caminho.name}"
                )

            registros = iter(leitor(caminho))
            try:
                primeiro = next(registros)
            except StopIteration:
                continue

            self._validar_colunas_dbf(primeiro)
            for registro in chain((primeiro,), registros):
                registros_encontrados += 1
                yield self._converter_registro_dbf(registro)

        if registros_encontrados == 0:
            raise ValueError(
                "Nenhum registro foi encontrado nos bancos DBF."
            )

    def processar(
        self,
        municipios: Sequence[MunicipioQualifica],
        notificacoes: Iterable[Notificacao72h],
        data_inicial: date,
        data_final: date,
    ) -> ResultadoRelatorio72h:
        data_inicial = self._normalizar_data_limite(
            data_inicial,
            "inicial",
        )
        data_final = self._normalizar_data_limite(
            data_final,
            "final",
        )
        if data_inicial > data_final:
            raise ValueError(
                "A data inicial é posterior à data final."
            )

        municipios_por_codigo = {
            municipio.codigo_ibge: municipio
            for municipio in municipios
        }
        if len(municipios_por_codigo) != len(municipios):
            raise ValueError(
                "A lista de municípios possui códigos IBGE duplicados."
            )

        contagens = {
            codigo: [0, 0]
            for codigo in municipios_por_codigo
        }
        numeros_notificacao: Counter[str] = Counter()
        fora_do_dicionario = 0

        for notificacao in notificacoes:
            data_sintomas = notificacao.data_primeiros_sintomas
            if (
                data_sintomas is None
                or data_sintomas < data_inicial
                or data_sintomas > data_final
            ):
                continue

            codigo = self.normalizar_codigo_ibge(
                notificacao.municipio_notificacao,
                permitir_vazio=True,
            )
            if codigo not in contagens:
                fora_do_dicionario += 1
                continue

            contagens[codigo][0] += 1
            numero = self.normalizar_numero_notificacao(
                notificacao.numero
            )
            if numero:
                numeros_notificacao[numero] += 1

            data_notificacao = notificacao.data_notificacao
            data_digitacao = notificacao.data_digitacao
            if data_notificacao is None or data_digitacao is None:
                continue

            dias = (data_digitacao - data_notificacao).days
            if 0 <= dias <= 3:
                contagens[codigo][1] += 1

        resultados: list[ResultadoMunicipio72h] = []
        for codigo, municipio in municipios_por_codigo.items():
            total, dentro_prazo = contagens[codigo]
            percentual = (
                round((dentro_prazo / total) * 100, 2)
                if total > 0
                else 0.0
            )
            resultados.append(
                ResultadoMunicipio72h(
                    codigo_ibge=codigo,
                    nome=municipio.nome,
                    crs=municipio.crs,
                    numero_notificacoes=total,
                    dias_dentro_do_prazo=dentro_prazo,
                    fora_prazo_72h=total - dentro_prazo,
                    percentual=percentual,
                )
            )

        resultados.sort(
            key=lambda item: (
                item.crs,
                self._texto_para_ordenacao(item.nome),
            )
        )
        total_notificacoes = sum(
            item.numero_notificacoes
            for item in resultados
        )
        total_dentro_prazo = sum(
            item.dias_dentro_do_prazo
            for item in resultados
        )
        percentual_estadual = (
            round(
                (total_dentro_prazo / total_notificacoes) * 100,
                2,
            )
            if total_notificacoes > 0
            else 0.0
        )

        avisos: list[str] = []
        repetidos = sum(
            1
            for quantidade in numeros_notificacao.values()
            if quantidade > 1
        )
        if repetidos:
            avisos.append(
                f"Foram encontrados {repetidos} número(s) de "
                "notificação repetido(s). As linhas foram mantidas."
            )
        if fora_do_dicionario:
            avisos.append(
                f"Foram desconsideradas {fora_do_dicionario} "
                "notificação(ões) com município ausente ou fora do "
                "dicionário."
            )

        return ResultadoRelatorio72h(
            data_inicial=data_inicial,
            data_final=data_final,
            municipios=tuple(resultados),
            total_notificacoes=total_notificacoes,
            total_dentro_do_prazo=total_dentro_prazo,
            percentual_estadual=percentual_estadual,
            avisos=tuple(avisos),
        )

    def exportar_excel(
        self,
        resultado: ResultadoRelatorio72h,
        caminho_saida: str | Path,
    ) -> Path:
        destino = Path(caminho_saida).expanduser()
        if destino.suffix.casefold() != ".xlsx":
            raise ValueError(
                "O relatório de 72h precisa ser salvo como XLSX."
            )
        destino.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        resumo = workbook.active
        resumo.title = "Resumo_Estadual"
        dados = workbook.create_sheet("Dados_Municipios")

        resumo.append(["Métrica", "Valor"])
        resumo.append(
            ["Total de Notificações", resultado.total_notificacoes]
        )
        resumo.append(
            [
                "Total Digitadas no Prazo",
                resultado.total_dentro_do_prazo,
            ]
        )
        resumo.append(
            [
                "Percentual de Oportunidade Estadual (%)",
                resultado.percentual_estadual,
            ]
        )

        dados.append(list(self.CABECALHOS_SAIDA))
        for municipio in resultado.municipios:
            dados.append(
                [
                    municipio.codigo_ibge,
                    municipio.nome,
                    municipio.crs,
                    municipio.numero_notificacoes,
                    municipio.dias_dentro_do_prazo,
                    municipio.fora_prazo_72h,
                    municipio.percentual,
                ]
            )

        self._formatar_planilha_resumo(resumo)
        self._formatar_planilha_municipios(dados)
        try:
            self._salvar_workbook_atomico(workbook, destino)
        finally:
            workbook.close()
        return destino

    def gerar_relatorio(
        self,
        caminho_dicionario: str | Path,
        caminhos_dbf: Sequence[str | Path],
        data_inicial: date,
        data_final: date,
        caminho_saida: str | Path,
        *,
        callback_status: Callable[[str], None] | None = None,
        leitor_dbf: Callable[
            [Path],
            Iterable[Mapping[str, Any]],
        ] | None = None,
    ) -> ResultadoRelatorio72h:
        destino = Path(caminho_saida).expanduser()
        origens = {
            Path(caminho_dicionario).expanduser().resolve(),
            *(
                Path(caminho).expanduser().resolve()
                for caminho in caminhos_dbf
            ),
        }
        if destino.resolve() in origens:
            raise ValueError(
                "O arquivo de saída não pode substituir um arquivo de "
                "entrada."
            )

        self._emitir_status(
            callback_status,
            "Validando o dicionário de municípios.",
        )
        municipios = self.carregar_municipios(caminho_dicionario)

        self._emitir_status(
            callback_status,
            f"Lendo {len(caminhos_dbf)} banco(s) DBF do SINAN.",
        )
        notificacoes = self.iterar_notificacoes_dbf(
            caminhos_dbf,
            leitor_dbf=leitor_dbf,
        )

        self._emitir_status(
            callback_status,
            "Calculando a oportunidade de digitação em 72 horas.",
        )
        resultado = self.processar(
            municipios=municipios,
            notificacoes=notificacoes,
            data_inicial=data_inicial,
            data_final=data_final,
        )

        self._emitir_status(
            callback_status,
            "Gerando o relatório consolidado em Excel.",
        )
        self.exportar_excel(resultado, destino)
        self._emitir_status(
            callback_status,
            "Relatório de 72h concluído.",
        )
        return resultado

    @staticmethod
    def normalizar_codigo_ibge(
        valor: Any,
        *,
        permitir_vazio: bool = False,
    ) -> str:
        if valor is None:
            if permitir_vazio:
                return ""
            raise ValueError("O código IBGE está vazio.")

        if isinstance(valor, float):
            if not math.isfinite(valor) or not valor.is_integer():
                raise ValueError(
                    f"Código IBGE inválido: {valor!r}."
                )
            texto = str(int(valor))
        else:
            texto = str(valor).strip()
            if re.fullmatch(r"\d+\.0", texto):
                texto = texto[:-2]

        if not texto:
            if permitir_vazio:
                return ""
            raise ValueError("O código IBGE está vazio.")
        if not re.fullmatch(r"\d{1,6}", texto):
            raise ValueError(
                f"Código IBGE inválido: {valor!r}."
            )
        return texto.zfill(6)

    @staticmethod
    def normalizar_numero_notificacao(valor: Any) -> str:
        if valor is None:
            return ""
        if isinstance(valor, float):
            if not math.isfinite(valor):
                return ""
            if valor.is_integer():
                return str(int(valor))

        texto = str(valor).strip()
        if re.fullmatch(r"\d+\.0", texto):
            return texto[:-2]
        return texto

    @staticmethod
    def converter_data(valor: Any) -> date | None:
        if valor in (None, ""):
            return None
        if isinstance(valor, datetime):
            return valor.date()
        if isinstance(valor, date):
            return valor

        texto = str(valor).strip()
        if not texto:
            return None
        for formato in (
            "%Y%m%d",
            "%d/%m/%Y",
            "%Y-%m-%d",
            "%Y-%m-%d %H:%M:%S",
        ):
            try:
                return datetime.strptime(texto, formato).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def _abrir_dbf(
        caminho: Path,
    ) -> Iterable[Mapping[str, Any]]:
        try:
            from dbfread import DBF
        except ImportError as erro:
            raise RuntimeError(
                "A dependência dbfread não está instalada. "
                "Reinstale as dependências do ArboHub."
            ) from erro

        return DBF(
            str(caminho),
            encoding="latin1",
            load=False,
            char_decode_errors="strict",
        )

    @classmethod
    def _converter_registro_dbf(
        cls,
        registro: Mapping[str, Any],
    ) -> Notificacao72h:
        return Notificacao72h(
            numero=cls.normalizar_numero_notificacao(
                registro.get("NU_NOTIFIC")
            ),
            municipio_notificacao=cls.normalizar_codigo_ibge(
                registro.get("ID_MUNICIP"),
                permitir_vazio=True,
            ),
            data_primeiros_sintomas=cls.converter_data(
                registro.get("DT_SIN_PRI")
            ),
            data_notificacao=cls.converter_data(
                registro.get("DT_NOTIFIC")
            ),
            data_digitacao=cls.converter_data(
                registro.get("DT_DIGITA")
            ),
        )

    @staticmethod
    def _validar_colunas_dbf(registro: Mapping[str, Any]):
        faltantes = sorted(
            COLUNAS_DBF_OBRIGATORIAS - set(registro)
        )
        if faltantes:
            raise ValueError(
                "O banco SINAN não contém as colunas necessárias: "
                + ", ".join(faltantes)
            )

    @staticmethod
    def _normalizar_cabecalho(valor: Any) -> str:
        texto = unicodedata.normalize(
            "NFKD",
            str(valor or "").strip(),
        )
        texto = "".join(
            caractere
            for caractere in texto
            if not unicodedata.combining(caractere)
        )
        texto = re.sub(r"[^a-z0-9]+", " ", texto.casefold())
        return " ".join(texto.split())

    @staticmethod
    def _texto_para_ordenacao(valor: str) -> str:
        texto = unicodedata.normalize("NFKD", valor.casefold())
        return "".join(
            caractere
            for caractere in texto
            if not unicodedata.combining(caractere)
        )

    @staticmethod
    def _valor_linha(linha: tuple, indice: int) -> Any:
        return linha[indice] if indice < len(linha) else None

    @staticmethod
    def _normalizar_crs(valor: Any, numero_linha: int) -> int:
        try:
            numero = float(valor)
        except (TypeError, ValueError) as erro:
            raise ValueError(
                f"CRS inválida no dicionário, linha {numero_linha}."
            ) from erro
        if not math.isfinite(numero) or not numero.is_integer():
            raise ValueError(
                f"CRS inválida no dicionário, linha {numero_linha}."
            )
        return int(numero)

    @staticmethod
    def _normalizar_data_limite(
        valor: date,
        descricao: str,
    ) -> date:
        data_convertida = Relatorio72hService.converter_data(valor)
        if data_convertida is None:
            raise ValueError(
                f"A data {descricao} do relatório é inválida."
            )
        return data_convertida

    @staticmethod
    def _emitir_status(
        callback: Callable[[str], None] | None,
        mensagem: str,
    ):
        if callback:
            callback(mensagem)

    @staticmethod
    def _formatar_cabecalho(planilha, intervalo: str):
        preenchimento = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )
        for celula in planilha[intervalo][0]:
            celula.fill = preenchimento
            celula.font = Font(
                color="FFFFFF",
                bold=True,
            )
            celula.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

    @classmethod
    def _formatar_planilha_resumo(cls, planilha):
        cls._formatar_cabecalho(planilha, "A1:B1")
        planilha.column_dimensions["A"].width = 44
        planilha.column_dimensions["B"].width = 18
        planilha["B4"].number_format = "0.00"
        planilha.freeze_panes = "A2"
        planilha.auto_filter.ref = "A1:B4"

    @classmethod
    def _formatar_planilha_municipios(cls, planilha):
        cls._formatar_cabecalho(planilha, "A1:G1")
        larguras = {
            "A": 14,
            "B": 34,
            "C": 8,
            "D": 24,
            "E": 24,
            "F": 18,
            "G": 14,
        }
        for coluna, largura in larguras.items():
            planilha.column_dimensions[coluna].width = largura
        for celula in planilha["A"][1:]:
            celula.number_format = "@"
        for celula in planilha["G"][1:]:
            celula.number_format = "0.00"
        planilha.freeze_panes = "A2"
        planilha.auto_filter.ref = planilha.dimensions

    @staticmethod
    def _salvar_workbook_atomico(
        workbook: Workbook,
        destino: Path,
    ):
        descritor, nome_temporario = tempfile.mkstemp(
            prefix=f".{destino.stem}_",
            suffix=".tmp.xlsx",
            dir=destino.parent,
        )
        os.close(descritor)
        temporario = Path(nome_temporario)

        try:
            workbook.save(temporario)
            verificacao = load_workbook(
                temporario,
                read_only=True,
                data_only=True,
            )
            try:
                esperadas = {
                    "Resumo_Estadual",
                    "Dados_Municipios",
                }
                if not esperadas.issubset(verificacao.sheetnames):
                    raise ValueError(
                        "O relatório temporário não contém as abas "
                        "esperadas."
                    )
            finally:
                verificacao.close()
            os.replace(temporario, destino)
        finally:
            temporario.unlink(missing_ok=True)
